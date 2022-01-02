from os import link
from openpyxl.descriptors.base import Length
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Log
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.expected_conditions import presence_of_element_located
from selenium.common.exceptions import NoSuchElementException


#Create Workbook and name Sheet 1
wk = openpyxl.Workbook()
sh = wk.active
sh.title = "Ebay All Int."

#Create Sheet 2
wk.create_sheet(title="Ebay Domestic")
sh1 = wk["Ebay Domestic"]

#Create Headers
sh["A1"].value="Title"
sh["B1"].value="Price"
sh["C1"].value="Link"

sh1.cell(1,1).value="Title"
sh1["B1"].value="Price"
sh1["C1"].value="Link"

wk.save("C:\\Users\Jerry\Documents\Python Automation Project\Odyssey Go1.xlsx")

#Direct website to info I want to import

driver = webdriver.Chrome()

driver.get("https://www.geo-ship.com/#/");

try:

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-app-topbar/div/div[1]/form/div[1]/p-autocomplete/span/input"))
    )
    element.send_keys("Magnavox Odyssey")

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-app-topbar/div/div[1]/form/div[3]/button/span[1]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[1]/p-checkbox/div"))
    )
    element.click()


    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[4]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[7]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[10]/p-checkbox/div/div[2]"))
    )
    element.click()
    
    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[13]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[2]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[5]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[8]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[11]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[3]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[6]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[9]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[1]/div/div/div[2]/div/div[12]/p-checkbox/div/div[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[2]/div/div/div/div[2]/div/p-dropdown/div/span"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[2]/div/div/div/div[2]/div/p-dropdown/div/div[3]/div/ul/p-dropdownitem[3]/li"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-advanced-search/div/div/p-card[2]/div/div/div/div[2]/div/p-dropdown/div/div[3]/div/ul/p-dropdownitem[3]/li"))
    )
    element.click()

    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/app-root/app-app-topbar/div/div[1]/form/div[1]/button/span[1]"))
    )
    element.click()

#Import International Ebay Listings

except:
    time.sleep(10)

    #Create Data List

WebDriverWait(driver,10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "title-link"))
)
game_titles = driver.find_elements_by_class_name( "title-link")
title_list = []
for i in game_titles:
    title_list.append(i.text)
title_iter = 1

while title_iter < len(title_list):
    print(title_list[title_iter])
    sh.cell(title_iter,1).value=title_list[title_iter]
    title_iter += 1

wk.save("C:\\Users\Jerry\Documents\Python Automation Project\Odyssey Go1.xlsx")


game_prices = driver.find_elements_by_class_name( "strong-price")
price_list = []
for x in game_prices:
    price_list.append(x.text)
price_iter = 1

while price_iter < len(price_list):
    print(price_list[price_iter])
    sh.cell(price_iter,2).value=price_list[price_iter]
    price_iter += 1

wk.save("C:\\Users\Jerry\Documents\Python Automation Project\Odyssey Go1.xlsx")


game_links = driver.find_elements_by_xpath('//div[@class="p-ml-1 p-mb-2 p-mt-1"]/a/@href') 
link_list = []
for y in game_links:
    link_list.append(y.text)
link_iter = 1

while link_iter < len(link_list):
    print(link_list[link_iter])
    sh.cell(link_iter,3).value=link_list[link_iter]
    link_iter += 1

wk.save("C:\\Users\Jerry\Documents\Python Automation Project\Odyssey Go1.xlsx")








time.sleep(30)




